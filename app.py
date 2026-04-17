"""
Inventário de Frota — Automação de Evidências
Lê PDFs de inventário, extrai placas/chassis via OCR e preenche
a coluna 'Evidencia' do XLSX com "SIM" para os ativos encontrados.
"""

import streamlit as st
import fitz  # PyMuPDF
from PIL import Image
import numpy as np
import re
import io
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pandas as pd

# ─────────────────────────────────────────────────────────────
# Configuração da página
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Inventário de Frota",
    page_icon="🚛",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
/* Oculta header, menu e footer padrão do Streamlit */
#MainMenu          { visibility: hidden; }
header             { visibility: hidden; }
footer             { visibility: hidden; }
[data-testid="stToolbar"] { display: none; }

/* Tipografia e cores gerais */
html, body, [class*="css"] {
    font-family: 'Segoe UI', 'Arial', sans-serif;
}
.main .block-container {
    padding-top: 1.8rem;
    max-width: 1100px;
}

/* Header da aplicação */
.app-header {
    background: linear-gradient(135deg, #1a237e 0%, #1565C0 100%);
    border-radius: 12px;
    padding: 1.5rem 2rem;
    color: #fff;
    margin-bottom: 1.5rem;
    display: flex;
    align-items: center;
    gap: 1rem;
}
.app-header h1 { margin: 0; font-size: 1.6rem; font-weight: 700; }
.app-header p  { margin: 0.2rem 0 0; opacity: 0.85; font-size: 0.95rem; }

/* Cards de upload */
.upload-card {
    background: #fff;
    border-radius: 10px;
    padding: 1.3rem 1.5rem;
    box-shadow: 0 1px 6px rgba(0,0,0,0.08);
    height: 100%;
}
.upload-card h3 { margin: 0 0 0.5rem; color: #1565C0; font-size: 1rem; }
.upload-card p  { color: #666; font-size: 0.83rem; margin: 0 0 0.8rem; }

/* Métricas customizadas */
.metric-row {
    display: flex;
    gap: 1rem;
    margin: 0.5rem 0 1rem;
}
.metric-box {
    flex: 1;
    background: #fff;
    border-radius: 10px;
    padding: 1.1rem 1rem;
    text-align: center;
    box-shadow: 0 1px 6px rgba(0,0,0,0.08);
}
.metric-box .number {
    font-size: 2rem;
    font-weight: 700;
    line-height: 1;
    color: #1565C0;
}
.metric-box .label {
    font-size: 0.78rem;
    color: #666;
    margin-top: 0.35rem;
    text-transform: uppercase;
    letter-spacing: 0.03em;
}
.metric-box.green  .number { color: #2e7d32; }
.metric-box.orange .number { color: #e65100; }

/* Seção de resultados */
.results-section {
    background: #fff;
    border-radius: 10px;
    padding: 1.5rem;
    box-shadow: 0 1px 6px rgba(0,0,0,0.08);
    margin-top: 1rem;
}
.results-section h3 { color: #1a237e; margin-top: 0; }

/* Botão primário override */
div[data-testid="stButton"] > button[kind="primary"] {
    background: #1565C0;
    border: none;
    font-weight: 600;
    letter-spacing: 0.02em;
    height: 3rem;
    font-size: 1rem;
}
div[data-testid="stButton"] > button[kind="primary"]:hover {
    background: #0d47a1;
}

/* Divider */
hr { border: none; border-top: 1px solid #e0e0e0; margin: 1.2rem 0; }

/* Status / spinner texto */
.stStatus { border-radius: 8px; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# OCR Engine — carregado uma vez, persistido em cache
# ─────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def carregar_ocr():
    """Inicializa EasyOCR — detecta GPU automaticamente."""
    import easyocr
    try:
        reader = easyocr.Reader(["en"], gpu=True, verbose=False)
        return reader, True
    except Exception:
        reader = easyocr.Reader(["en"], gpu=False, verbose=False)
        return reader, False


# ─────────────────────────────────────────────────────────────
# Regex — padrões de placas brasileiras
# ─────────────────────────────────────────────────────────────
PAT_MERCOSUL = re.compile(r"[A-Z]{3}[0-9][A-Z][0-9]{2}")   # ABC1D23
PAT_ANTIGO   = re.compile(r"[A-Z]{3}[0-9]{4}")              # ABC1234
PLATE_PATTERNS = [PAT_MERCOSUL, PAT_ANTIGO]


def normalizar(texto: str) -> str:
    """Remove tudo que não é letra/número e converte para maiúsculas."""
    return re.sub(r"[^A-Z0-9]", "", str(texto).upper())


# ─────────────────────────────────────────────────────────────
# Extração de imagens dos PDFs
# ─────────────────────────────────────────────────────────────
def extrair_imagens_pdf(pdf_files) -> list:
    """
    Extrai todas as imagens embutidas dos PDFs enviados.
    Fallback: se uma página não tiver imagens embutidas,
    renderiza a página inteira (2× zoom ≈ 144 dpi).
    """
    imagens = []
    for pdf_file in pdf_files:
        nome_pdf = pdf_file.name
        try:
            pdf_bytes = pdf_file.read()
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        except Exception:
            continue

        for pag_idx in range(len(doc)):
            pagina = doc[pag_idx]
            lista_imgs = pagina.get_images(full=True)
            adicionadas = 0

            for img_info in lista_imgs:
                xref = img_info[0]
                try:
                    base = doc.extract_image(xref)
                    pil_img = Image.open(io.BytesIO(base["image"])).convert("RGB")
                    w, h = pil_img.size
                    if w < 80 or h < 80:        # descarta miniaturas
                        continue
                    imagens.append({
                        "pil":  pil_img,
                        "raw":  base["image"],
                        "ext":  base.get("ext", "jpg"),
                        "src":  nome_pdf,
                        "pag":  pag_idx + 1,
                        "seq":  len(imagens),
                    })
                    adicionadas += 1
                except Exception:
                    pass

            # Fallback: renderiza a página se não encontrou imagens embutidas
            if adicionadas == 0:
                try:
                    mat = fitz.Matrix(2, 2)
                    pix = pagina.get_pixmap(matrix=mat)
                    pil_img = Image.frombytes("RGB",
                                              [pix.width, pix.height],
                                              pix.samples)
                    raw_jpg = io.BytesIO()
                    pil_img.save(raw_jpg, format="JPEG", quality=90)
                    imagens.append({
                        "pil":  pil_img,
                        "raw":  raw_jpg.getvalue(),
                        "ext":  "jpg",
                        "src":  nome_pdf,
                        "pag":  pag_idx + 1,
                        "seq":  len(imagens),
                    })
                except Exception:
                    pass

    return imagens


# ─────────────────────────────────────────────────────────────
# Construção do lookup da planilha
# ─────────────────────────────────────────────────────────────
def construir_lookup(ws):
    """
    Retorna dois dicts:
      - placa_lookup:  {placa_normalizada: linha_xlsx}
      - chassis_lookup: {chassi_normalizado: linha_xlsx}
    """
    placa_lkp   = {}
    chassis_lkp = {}

    for linha in range(4, ws.max_row + 1):
        val_a = ws.cell(linha, 1).value   # Placa LW
        val_b = ws.cell(linha, 2).value   # Placa AX
        val_d = ws.cell(linha, 4).value   # Chassi

        for v in [val_a, val_b]:
            if v:
                n = normalizar(str(v))
                if n:
                    placa_lkp[n] = linha

        if val_d:
            n = normalizar(str(val_d))
            if len(n) >= 5:
                chassis_lkp[n] = linha

    return placa_lkp, chassis_lkp


# ─────────────────────────────────────────────────────────────
# OCR + busca de candidatos
# ─────────────────────────────────────────────────────────────
def ocr_e_candidatos(pil_img, reader, conf_min: float = 0.30):
    """
    Executa OCR na imagem e retorna:
      - set de candidatos (placas + sequências longas para chassi)
      - string concatenada de todo o texto reconhecido (para busca substring)
    """
    arr = np.array(pil_img)
    resultados = reader.readtext(arr, detail=1, paragraph=False)

    textos_norm = [
        normalizar(r[1])
        for r in resultados
        if r[2] >= conf_min and len(r[1].strip()) >= 2
    ]

    candidatos = set()
    concat_total = "".join(textos_norm)   # sem espaços — para regex e substring

    # Busca de placas em segmentos individuais e pares adjacentes
    for i, seg in enumerate(textos_norm):
        for pat in PLATE_PATTERNS:
            for m in pat.finditer(seg):
                candidatos.add(m.group())

        if i < len(textos_norm) - 1:
            par = seg + textos_norm[i + 1]
            for pat in PLATE_PATTERNS:
                for m in pat.finditer(par):
                    candidatos.add(m.group())

    # Sequências longas como potenciais chassi
    for seg in textos_norm:
        if len(seg) >= 10:
            candidatos.add(seg)

    return candidatos, concat_total


# ─────────────────────────────────────────────────────────────
# Correspondência com a planilha
# ─────────────────────────────────────────────────────────────
def buscar_correspondencia(candidatos, concat_total, placa_lkp, chassis_lkp):
    """
    Retorna (identificador_encontrado, linha_xlsx) ou (None, None).
    Prioridade: Placa LW/AX → depois Chassi.
    """
    # 1. Placas (Mercosul e antigas)
    for c in candidatos:
        if c in placa_lkp:
            return c, placa_lkp[c]

    # 2. Chassi: busca substring do valor conhecido no texto OCR
    for chassi_norm, linha in chassis_lkp.items():
        if len(chassi_norm) >= 10 and chassi_norm in concat_total:
            return chassi_norm, linha

    return None, None


# ─────────────────────────────────────────────────────────────
# Atualização da planilha — APENAS coluna M (Evidencia)
# ─────────────────────────────────────────────────────────────
def atualizar_xlsx(xlsx_bytes: bytes, linhas_encontradas: set) -> bytes:
    """
    Carrega o XLSX, limpa a coluna Evidencia e preenche "SIM"
    somente nas linhas identificadas. Retorna bytes do XLSX atualizado.
    Preserva toda a formatação, fórmulas e estrutura da planilha.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Inventário"]

    alinhamento_sim = Alignment(horizontal="center")

    for linha in range(4, ws.max_row + 1):
        if not ws.cell(linha, 1).value:
            continue
        celula = ws.cell(linha, 13)   # coluna M = Evidencia
        if linha in linhas_encontradas:
            celula.value     = "SIM"
            celula.alignment = alinhamento_sim
        else:
            celula.value = None

    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()


# ─────────────────────────────────────────────────────────────
# Construção do ZIP com imagens não identificadas
# ─────────────────────────────────────────────────────────────
def montar_zip_nao_identificadas(imagens: list) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for i, img in enumerate(imagens):
            nome_pdf = img["src"].replace(".pdf", "").replace(" ", "_")
            nome_arq = f"{nome_pdf}_pag{img['pag']}_{i:04d}.{img['ext']}"
            zf.writestr(nome_arq, img["raw"])
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# Interface
# ─────────────────────────────────────────────────────────────

st.markdown("""
<div class="app-header">
  <div>
    <h1>🚛 Inventário de Frota</h1>
    <p>Automação de evidências — leitura de placas e chassis via OCR nos PDFs do inventário</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Uploads ──────────────────────────────────────────────────
col_pdf, col_xlsx = st.columns(2, gap="large")

with col_pdf:
    st.markdown('<div class="upload-card"><h3>📄 PDFs do Inventário</h3>'
                '<p>Selecione um ou mais arquivos PDF gerados no inventário de frota</p></div>',
                unsafe_allow_html=True)
    pdf_uploads = st.file_uploader(
        "PDFs",
        type=["pdf"],
        accept_multiple_files=True,
        label_visibility="collapsed",
        key="pdf_uploader",
    )
    if pdf_uploads:
        st.caption(f"✅ {len(pdf_uploads)} arquivo(s) selecionado(s): "
                   + ", ".join(f.name for f in pdf_uploads[:3])
                   + ("…" if len(pdf_uploads) > 3 else ""))

with col_xlsx:
    st.markdown('<div class="upload-card"><h3>📊 Planilha XLSX</h3>'
                '<p>Planilha de inventário sem a coluna Evidência preenchida</p></div>',
                unsafe_allow_html=True)
    xlsx_upload = st.file_uploader(
        "XLSX",
        type=["xlsx"],
        accept_multiple_files=False,
        label_visibility="collapsed",
        key="xlsx_uploader",
    )
    if xlsx_upload:
        st.caption(f"✅ Arquivo selecionado: {xlsx_upload.name}")

st.markdown("<hr>", unsafe_allow_html=True)

pronto = bool(pdf_uploads and xlsx_upload)

btn_processar = st.button(
    "▶  Processar Inventário",
    type="primary",
    use_container_width=True,
    disabled=not pronto,
)

if not pronto:
    st.info("⬆️ Envie os PDFs e a planilha XLSX acima para habilitar o processamento.")


# ─────────────────────────────────────────────────────────────
# Processamento principal
# ─────────────────────────────────────────────────────────────
if btn_processar and pronto:

    # ── 0. Inicializar OCR ────────────────────────────────────
    with st.spinner("🔧 Inicializando motor de OCR (primeira execução pode demorar)…"):
        reader, usando_gpu = carregar_ocr()
    modo_hw = "🖥️ GPU (acelerado)" if usando_gpu else "💻 CPU"
    st.caption(f"Motor OCR pronto — {modo_hw}")

    # ── 1. Carregar XLSX ──────────────────────────────────────
    with st.status("📂 Carregando planilha…", expanded=False) as status_xlsx:
        xlsx_bytes = xlsx_upload.read()
        wb_tmp = load_workbook(io.BytesIO(xlsx_bytes))
        ws_tmp = wb_tmp["Inventário"]
        placa_lkp, chassis_lkp = construir_lookup(ws_tmp)
        total_registros = sum(
            1 for r in range(4, ws_tmp.max_row + 1)
            if ws_tmp.cell(r, 1).value
        )
        total_ids = len(placa_lkp) + len(chassis_lkp)
        status_xlsx.update(
            label=f"✅ Planilha carregada — {total_registros} registros | {total_ids} identificadores indexados",
            state="complete",
        )

    # ── 2. Extrair imagens dos PDFs ───────────────────────────
    with st.status("🖼️ Extraindo imagens dos PDFs…", expanded=False) as status_imgs:
        # Rewind file pointers
        for f in pdf_uploads:
            f.seek(0)
        imagens = extrair_imagens_pdf(pdf_uploads)
        status_imgs.update(
            label=f"✅ {len(imagens)} imagens extraídas de {len(pdf_uploads)} PDF(s)",
            state="complete",
        )

    if not imagens:
        st.error("❌ Nenhuma imagem foi extraída dos PDFs enviados. Verifique os arquivos.")
        st.stop()

    # ── 3. OCR + Correspondência ──────────────────────────────
    st.markdown("#### 🔍 Processando imagens…")
    barra   = st.progress(0.0)
    info_st = st.empty()

    linhas_encontradas = set()
    nao_identificadas  = []
    log_matches        = []

    for i, img_data in enumerate(imagens):
        info_st.caption(
            f"Imagem {i+1}/{len(imagens)} — "
            f"{img_data['src']} · pág. {img_data['pag']}"
        )
        try:
            candidatos, concat = ocr_e_candidatos(img_data["pil"], reader)
            id_match, linha_match = buscar_correspondencia(
                candidatos, concat, placa_lkp, chassis_lkp
            )
            if linha_match:
                linhas_encontradas.add(linha_match)
                log_matches.append({
                    "Identificador": id_match,
                    "Linha XLSX": linha_match,
                    "PDF": img_data["src"],
                    "Página": img_data["pag"],
                })
            else:
                nao_identificadas.append(img_data)
        except Exception:
            nao_identificadas.append(img_data)

        barra.progress((i + 1) / len(imagens))

    barra.empty()
    info_st.empty()

    # ── 4. Atualizar e salvar XLSX ────────────────────────────
    with st.status("💾 Atualizando coluna Evidência…", expanded=False) as status_save:
        xlsx_atualizado = atualizar_xlsx(xlsx_bytes, linhas_encontradas)
        status_save.update(
            label=f"✅ Coluna Evidência atualizada — {len(linhas_encontradas)} linhas marcadas como SIM",
            state="complete",
        )

    # ── 5. Montar ZIP com não identificadas ──────────────────
    zip_nao_id = None
    if nao_identificadas:
        with st.status("📦 Preparando ZIP das imagens não identificadas…", expanded=False) as st_zip:
            zip_nao_id = montar_zip_nao_identificadas(nao_identificadas)
            st_zip.update(
                label=f"✅ ZIP preparado com {len(nao_identificadas)} imagens",
                state="complete",
            )

    # ── Persistir no session_state ────────────────────────────
    st.session_state.update({
        "xlsx_atualizado":    xlsx_atualizado,
        "zip_nao_id":         zip_nao_id,
        "n_imagens":          len(imagens),
        "n_encontradas":      len(linhas_encontradas),
        "n_nao_identificadas": len(nao_identificadas),
        "log_matches":        log_matches,
    })
    st.success("✅ Processamento concluído com sucesso!")


# ─────────────────────────────────────────────────────────────
# Resultados e downloads
# ─────────────────────────────────────────────────────────────
if "xlsx_atualizado" in st.session_state:
    st.markdown("<hr>", unsafe_allow_html=True)

    ni  = st.session_state["n_imagens"]
    ne  = st.session_state["n_encontradas"]
    nni = st.session_state["n_nao_identificadas"]

    st.markdown(f"""
    <div class="metric-row">
      <div class="metric-box">
        <div class="number">{ni}</div>
        <div class="label">Imagens processadas</div>
      </div>
      <div class="metric-box green">
        <div class="number">{ne}</div>
        <div class="label">Registros com SIM</div>
      </div>
      <div class="metric-box {'orange' if nni else 'green'}">
        <div class="number">{nni}</div>
        <div class="label">Imagens não identificadas</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<hr>", unsafe_allow_html=True)

    # Downloads
    dl_col1, dl_col2 = st.columns(2, gap="large")

    with dl_col1:
        st.download_button(
            label="📥  Download Planilha Preenchida (.xlsx)",
            data=st.session_state["xlsx_atualizado"],
            file_name="INVENTARIO_FROTA_EVIDENCIAS.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

    with dl_col2:
        if st.session_state["zip_nao_id"]:
            st.download_button(
                label="📦  Download Imagens Não Identificadas (.zip)",
                data=st.session_state["zip_nao_id"],
                file_name="imagens_nao_identificadas.zip",
                mime="application/zip",
                use_container_width=True,
            )
        else:
            st.success("🎉 Todas as imagens foram identificadas na planilha!")

    # Log de correspondências
    if st.session_state.get("log_matches"):
        with st.expander(
            f"🔎 Ver detalhes dos {st.session_state['n_encontradas']} registros identificados"
        ):
            df_log = pd.DataFrame(st.session_state["log_matches"])
            st.dataframe(df_log, use_container_width=True, hide_index=True)
